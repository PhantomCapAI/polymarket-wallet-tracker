import asyncio
import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, text

from app.models.wallet import WalletMaster
from app.models.trade import TradesLog
from app.models.alert import AlertsLog
from app.models.copy_trade import CopyTrades
from app.models.market import MarketSummary
import logging

logger = logging.getLogger(__name__)

class ExcelExportService:
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.center_alignment = Alignment(horizontal="center")
    
    async def export_all_tables(self, db: AsyncSession) -> bytes:
        """Export all tables to Excel file"""
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Export each table
            await self._export_wallets_master(db, wb)
            await self._export_trades_log(db, wb)
            await self._export_market_summary(db, wb)
            await self._export_alerts_log(db, wb)
            await self._export_copy_trades(db, wb)
            
            # Save to bytes
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    async def _export_wallets_master(self, db: AsyncSession, wb: Workbook):
        """Export wallets_master table"""
        try:
            # Get data
            query = select(WalletMaster).order_by(WalletMaster.signal_score.desc())
            result = await db.execute(query)
            wallets = result.scalars().all()
            
            if not wallets:
                return
            
            # Convert to DataFrame
            wallet_data = []
            for wallet in wallets:
                wallet_data.append({
                    'Wallet': wallet.wallet,
                    'Signal Score': float(wallet.signal_score),
                    'Realized PnL': float(wallet.realized_pnl),
                    'Win Rate': float(wallet.win_rate),
                    'Avg Position Size': float(wallet.avg_position_size),
                    'Market Diversity': wallet.market_diversity,
                    'Timing Edge': wallet.timing_edge,
                    'Closing Efficiency': float(wallet.closing_efficiency),
                    'Consistency Score': float(wallet.consistency_score),
                    'Total Trades': wallet.total_trades,
                    'Active Days': wallet.active_days,
                    'Last Trade At': wallet.last_trade_at,
                    'Last Updated': wallet.last_updated
                })
            
            df = pd.DataFrame(wallet_data)
            
            # Create worksheet
            ws = wb.create_sheet("Wallets Master")
            
            # Add data
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            
            # Format headers
            self._format_headers(ws)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            logger.info(f"Exported {len(wallet_data)} wallets to Excel")
            
        except Exception as e:
            logger.error(f"Error exporting wallets_master: {e}")
    
    async def _export_trades_log(self, db: AsyncSession, wb: Workbook):
        """Export trades_log table"""
        try:
            # Get data (limit to recent trades for performance)
            query = select(TradesLog).order_by(
                TradesLog.entry_time.desc()
            ).limit(10000)
            result = await db.execute(query)
            trades = result.scalars().all()
            
            if not trades:
                return
            
            # Convert to DataFrame
            trade_data = []
            for trade in trades:
                trade_data.append({
                    'Trade ID': trade.id,
                    'Wallet': trade.wallet,
                    'Market': trade.market,
                    'Entry Time': trade.entry_time,
                    'Exit Time': trade.exit_time,
                    'Entry Price': float(trade.entry_price),
                    'Exit Price': float(trade.exit_price) if trade.exit_price else None,
                    'Peak Price': float(trade.peak_price) if trade.peak_price else None,
                    'Position Size': float(trade.position_size),
                    'PnL': float(trade.pnl) if trade.pnl else None,
                    'Outcome': trade.outcome,
                    'Fees': float(trade.fees) if trade.fees else 0.0,
                    'Created At': trade.created_at
                })
            
            df = pd.DataFrame(trade_data)
            
            # Create worksheet
            ws = wb.create_sheet("Trades Log")
            
            # Add data
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            
            # Format headers
            self._format_headers(ws)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            logger.info(f"Exported {len(trade_data)} trades to Excel")
            
        except Exception as e:
            logger.error(f"Error exporting trades_log: {e}")
    
    async def _export_market_summary(self, db: AsyncSession, wb: Workbook):
        """Export market_summary table"""
        try:
            query = select(MarketSummary).order_by(MarketSummary.total_volume.desc())
            result = await db.execute(query)
            markets = result.scalars().all()
            
            if not markets:
                return
            
            # Convert to DataFrame
            market_data = []
            for market in markets:
                market_data.append({
                    'Market': market.market,
                    'Total Volume': float(market.total_volume),
                    'Avg Win Rate': float(market.avg_win_rate),
                    'Top Wallet': market.top_wallet,
                    'Volatility': float(market.volatility),
                    'Trend Bias': market.trend_bias,
                    'Smart Money Count': market.smart_money_count,
                    'Last Updated': market.last_updated
                })
            
            df = pd.DataFrame(market_data)
            
            # Create worksheet
            ws = wb.create_sheet("Market Summary")
            
            # Add data
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            
            # Format headers
            self._format_headers(ws)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            logger.info(f"Exported {len(market_data)} markets to Excel")
            
        except Exception as e:
            logger.error(f"Error exporting market_summary: {e}")
    
    async def _export_alerts_log(self, db: AsyncSession, wb: Workbook):
        """Export alerts_log table"""
        try:
            # Get recent alerts (last 30 days)
            query = select(AlertsLog).order_by(
                AlertsLog.timestamp.desc()
            ).limit(5000)
            result = await db.execute(query)
            alerts = result.scalars().all()
            
            if not alerts:
                return
            
            # Convert to DataFrame
            alert_data = []
            for alert in alerts:
                alert_data.append({
                    'Alert ID': alert.id,
                    'Timestamp': alert.timestamp,
                    'Wallet': alert.wallet,
                    'Event Type': alert.event_type,
                    'Confidence': alert.confidence,
                    'Signal Reason': alert.signal_reason,
                    'Market': alert.market,
                    'Processed': alert.processed
                })
            
            df = pd.DataFrame(alert_data)
            
            # Create worksheet
            ws = wb.create_sheet("Alerts Log")
            
            # Add data
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            
            # Format headers
            self._format_headers(ws)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            logger.info(f"Exported {len(alert_data)} alerts to Excel")
            
        except Exception as e:
            logger.error(f"Error exporting alerts_log: {e}")
    
    async def _export_copy_trades(self, db: AsyncSession, wb: Workbook):
        """Export copy_trades table"""
        try:
            query = select(CopyTrades).order_by(CopyTrades.created_at.desc())
            result = await db.execute(query)
            copy_trades = result.scalars().all()
            
            if not copy_trades:
                return
            
            # Convert to DataFrame
            copy_trade_data = []
            for trade in copy_trades:
                copy_trade_data.append({
                    'Trade ID': trade.id,
                    'Source Wallet': trade.source_wallet,
                    'Market': trade.market,
                    'Direction': trade.direction,
                    'Entry Price': float(trade.entry_price),
                    'Exit Price': float(trade.exit_price) if trade.exit_price else None,
                    'Position Size': float(trade.position_size),
                    'Signal Score': float(trade.signal_score),
                    'Status': trade.status,
                    'PnL': float(trade.pnl) if trade.pnl else None,
                    'Stop Loss Price': float(trade.stop_loss_price) if trade.stop_loss_price else None,
                    'Fees': float(trade.fees) if trade.fees else 0.0,
                    'Execution Time': trade.execution_timestamp,
                    'Created At': trade.created_at,
                    'Closed At': trade.closed_at
                })
            
            df = pd.DataFrame(copy_trade_data)
            
            # Create worksheet
            ws = wb.create_sheet("Copy Trades")
            
            # Add data
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)
            
            # Format headers
            self._format_headers(ws)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            logger.info(f"Exported {len(copy_trade_data)} copy trades to Excel")
            
        except Exception as e:
            logger.error(f"Error exporting copy_trades: {e}")
    
    def _format_headers(self, ws):
        """Format worksheet headers"""
        for cell in ws[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

# Global export service instance
excel_export_service = ExcelExportService()
